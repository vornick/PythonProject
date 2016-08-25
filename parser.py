import sys, time, os, re, fileinput, csv, openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.cell import get_column_letter

def initialize_stat(ParsLogResultDir, num_of_platform):
    global EndResultDir, tcp, udp, l2, l3, mse, thr, con, pf, reg, st, con1, con16, con100, pac1, pac2, pac3, pac4, pac5, pac6, res_filename, file, i, count_Of_Zero
    EndResultDir = ParsLogResultDir
    tcp = 'TCP'
    udp='UDP'
    l2='L2'
    l3='L3'
    mse='CGW_MCE'
    thr='MaxThroughput'
    con='ConcurrentConnections'
    pf='25\\'
    reg='Reg'
    st='Standart'
    con1='1\\'
    con16='16\\'
    con100='100\\'
    pac1='256\\'
    pac2='512\\'
    pac3='1024\\'
    pac4='1280\\'
    pac5='1518\\'
    pac6='1400\\'
    i=0
    count_Of_Zero = 0
    res_filename = num_of_platform + '_TestResult.xlsx'
    file=EndResultDir + '\\' + res_filename
    print("Запускается сбор статистики в файл" + file)
    main(count_Of_Zero)
    return file

def main(count_Of_Zero):
    if not os.path.exists(EndResultDir):
        os.mkdir(EndResultDir)
    wb = Workbook()
    wb.save(file)
    global i
    tree = os.walk(EndResultDir)
    for d, dirs, files in tree:
        for f in files:
            if f == "HTTP_Client.csv":
                path=os.path.join (d,f)
                print(path)

                if path.find(tcp)>0:
                    u=1
                    write('A%s' % u, tcp)
                elif path.find(udp)>0:
                    u=30
                    write('A%s' % u, udp)

                if path.find(thr)>0:
                    a=u+1
                    write('B%s' % a, thr)
                elif path.find(con)>0:
                    a=u+15
                    write('B%s' % a, con)

                if path.find(l3)>0:
                    b=a+1
                    write('C%s' % b, 'L3')
                elif path.find(mse)>0:
                    b=a+5
                    write('C%s' % b, 'МСЭ')
                elif path.find(l2)>0:
                    b=a+9
                    write('C%s' % b, 'L2')

                if path.find(st)>0:
                    c=b+1
                    write('D%s' % c, 'Без регистрации')
                elif path.find(reg)>0:
                    c=b+2
                    write('D%s' % c, 'С регистрацией')
                elif path.find(pf)>0:
                    c=b+3
                    write('D%s' % c, '25 ПФ')

                if path.find(con1)>0:
                    c=b+1
                    write('D%s' % c, 1)
                elif path.find(con16)>0:
                    c=b+2
                    write('D%s' % c, 16)
                elif path.find(con100)>0:
                    c=b+3
                    write('D%s' % c, 100)

                if path.find(pac1)>0:
                    write('E%s' % a, 256)
                    csvfile(path)
                    write('E%s' % c, csvfile(path))
                elif path.find(pac2)>0:
                    write('F%s' % a, 512)
                    csvfile(path)
                    write('F%s' % c, csvfile(path))
                elif path.find(pac3)>0:
                    write('G%s' % a, 1024)
                    csvfile(path)
                    write('G%s' % c, csvfile(path))
                elif path.find(pac4)>0:
                    write('H%s' % a, 1280)
                    csvfile(path)
                    write('H%s' % c, csvfile(path))
                elif path.find(pac5)>0:
                    write('I%s' % a, 1518)
                    csvfile(path)
                    write('I%s' % c, csvfile(path))
                elif path.find(pac6)>0:
                    write('J%s' % a, 1400)
                    csvfile(path)
                    write('J%s' % c, csvfile(path))
                i+=1
                print('-----------------------------',i)

def csvfile(path):
    infile=open(path, 'r')
    rdr=csv.reader(infile)
    zn=0
    count=0
    zero_count=0
    srzn=0
    list_rdr = list(rdr)
    first_count = 16
    last_count = 76
    for cnt in range(first_count, last_count):
        row = list_rdr[cnt]
#		print(row)
        if int(row[95])>0:
            zn+=int(row[95])
            count+=1
            srzn=zn/count
        else:
            zero_count+=1
            #print(zero_count)
    infile.close()
    if zero_count==0:
        print('Passed')
    else:
        print('Fail',zero_count)
        zero_counter(zero_count)
    return srzn

def write(pos, val):
    wb = load_workbook(file)
    ws = wb.active
    ws[pos]=val
    wb.save(file)

def zero_counter(zero_count):
    global count_Of_Zero
    count_Of_Zero = count_Of_Zero + zero_count
    return count_Of_Zero